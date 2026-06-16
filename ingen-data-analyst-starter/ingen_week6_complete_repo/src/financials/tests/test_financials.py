"""Tests for the Week 6 peer-benchmark build (offline; uses sourced anchors, no network).

Run:  python -m pytest src/financials/tests -q
"""
from __future__ import annotations
from pathlib import Path
import openpyxl
import pytest
from src.financials import peer_data, build_workbook  # noqa: F401  (build runs on import)

ROOT = Path(__file__).resolve().parents[3]
WB = ROOT / "reports" / "week06" / "peer_financial_workbook.xlsx"


def test_workbook_has_required_sheets():
    wb = openpyxl.load_workbook(WB)
    for s in ["public_financials", "private_funding", "valuation_comps", "summary"]:
        assert s in wb.sheetnames, f"missing sheet {s}"


def test_every_public_row_has_source_and_sane_margins():
    for p in peer_data.PUBLIC:
        assert p["source"], f"{p['ticker']} missing source"
        assert -1 < p["gross_margin"] <= 1
        assert p["revenue"] > 0 and p["revenue_prior"] > 0


def test_every_private_round_cites_a_source():
    # success criterion: every private funding row cites a public source
    for p in peer_data.PRIVATE:
        assert p["source"] and len(p["source"]) > 8, f"{p['name']} round missing source"
        assert p["date"] and "-" in p["date"]


def test_revenue_growth_reconciles():
    # YoY computed in workbook should match a direct recompute within rounding
    wb = openpyxl.load_workbook(WB, data_only=True)
    ws = wb["public_financials"]
    # header row 2; find Revenue and Rev YoY columns
    hdr = {ws.cell(row=2, column=c).value: c for c in range(1, ws.max_column + 1)}
    for r in range(3, 3 + len(peer_data.PUBLIC)):
        name = ws.cell(row=r, column=hdr["Company"]).value
        rec = next(p for p in peer_data.PUBLIC if p["name"] == name)
        expected = rec["revenue"] / rec["revenue_prior"] - 1
        got = ws.cell(row=r, column=hdr["Rev YoY"]).value
        assert abs(got - expected) < 1e-6


def test_funding_per_employee_positive():
    # at least the well-funded humanoid privates have a derivable funding-per-employee
    for p in peer_data.PRIVATE:
        if p["amount_musd"] and p.get("employees"):
            assert p["amount_musd"] / p["employees"] > 0
