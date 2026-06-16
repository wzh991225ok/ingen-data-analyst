"""Build the Week 6 peer financial workbook (.xlsx) with four sheets:
public_financials, private_funding, valuation_comps, summary.

Run:  python -m src.financials.build_workbook
Output: reports/week06/peer_financial_workbook.xlsx
"""
from __future__ import annotations
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter as gl
from .peer_data import PUBLIC, PRIVATE, RETRIEVED

ROOT = Path(__file__).resolve().parents[2]
OUT = ROOT / "reports" / "week06" / "peer_financial_workbook.xlsx"
OUT.parent.mkdir(parents=True, exist_ok=True)

NAVY = "1F3864"; BLUE = "2E5496"; ALT = "F2F6FC"; YEL = "FFF2CC"
HF = Font(name="Calibri", bold=True, color="FFFFFF", size=10)
TF = Font(name="Calibri", bold=True, color="FFFFFF", size=13)
CF = Font(name="Calibri", size=10)
thin = Side(style="thin", color="D9D9D9"); BD = Border(left=thin, right=thin, top=thin, bottom=thin)


def _title(ws, text, ncols):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    c = ws.cell(row=1, column=1, value=text); c.font = TF
    c.fill = PatternFill("solid", fgColor=NAVY); c.alignment = Alignment(vertical="center", indent=1)
    ws.row_dimensions[1].height = 26


def _hdr(ws, headers, row=2):
    for c, h in enumerate(headers, 1):
        x = ws.cell(row=row, column=c, value=h); x.font = HF
        x.fill = PatternFill("solid", fgColor=BLUE); x.border = BD
        x.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[row].height = 30


def _rows(ws, data, start=3, pct_cols=(), flag_col=None):
    for i, r in enumerate(data):
        rr = start + i
        for c, v in enumerate(r, 1):
            x = ws.cell(row=rr, column=c, value=v); x.font = CF; x.border = BD
            x.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            if c in pct_cols and isinstance(v, (int, float)):
                x.number_format = "0.0%"
            if i % 2 == 1:
                x.fill = PatternFill("solid", fgColor=ALT)
            if flag_col and c == flag_col and isinstance(v, str) and ("approx" in v.lower() or "annualis" in v.lower() or "implied" in v.lower()):
                x.fill = PatternFill("solid", fgColor=YEL)


wb = Workbook(); wb.remove(wb.active)

# ---------------- public_financials ----------------
ws = wb.create_sheet("public_financials"); ws.sheet_view.showGridLines = False
_title(ws, "Public Peer Financials — FY2024 (SEC filings / earnings)", 9)
_hdr(ws, ["Ticker", "Company", "FY", "Revenue ($M)", "Rev YoY", "Gross margin",
          "R&D ($M)", "R&D % rev", "Op margin", ])  # note: source row below
# we put source + net margin in extra columns; redo header to 10 cols
ws.delete_rows(2)
_hdr(ws, ["Ticker", "Company", "FY", "Revenue ($M)", "Rev YoY", "Gross margin",
          "R&D ($M)", "R&D % rev", "Op margin", "Net margin", "Data quality", "Source"])
data = []
for p in PUBLIC:
    yoy = (p["revenue"] / p["revenue_prior"] - 1) if p["revenue_prior"] else None
    rnd_pct = p["rnd"] / p["revenue"] if p["revenue"] else None
    data.append([p["ticker"], p["name"], p["fy"], p["revenue"], yoy, p["gross_margin"],
                 p["rnd"], rnd_pct, p["op_margin"], p["net_margin"], p["data_quality"], p["source"]])
_rows(ws, data, pct_cols=(5, 6, 8, 9, 10), flag_col=11)
for c, w in enumerate([8, 12, 9, 12, 9, 11, 10, 10, 10, 10, 34, 52], 1):
    ws.column_dimensions[gl(c)].width = w
ws.freeze_panes = "A3"

# ---------------- private_funding ----------------
ws = wb.create_sheet("private_funding"); ws.sheet_view.showGridLines = False
_title(ws, "Private Peer Funding — disclosed rounds (press-sourced)", 8)
_hdr(ws, ["Company", "Round", "Date", "Amount ($M)", "Post-money ($B)",
          "Lead investors", "Employees", "Source"])
data = [[p["name"], p["round"], p["date"], p["amount_musd"], p["post_money_busd"],
         p["lead_investors"], p["employees"], p["source"]] for p in PRIVATE]
_rows(ws, data, flag_col=8)
for c, w in enumerate([16, 16, 9, 12, 14, 40, 11, 52], 1):
    ws.column_dimensions[gl(c)].width = w
ws.freeze_panes = "A3"

# ---------------- valuation_comps ----------------
ws = wb.create_sheet("valuation_comps"); ws.sheet_view.showGridLines = False
_title(ws, "Valuation Comparables — public growth/intensity & private capital efficiency", 7)
_hdr(ws, ["Company", "Type", "Rev growth / scale", "Gross margin", "R&D intensity",
          "Capital efficiency", "Caveat"])
comp = []
for p in PUBLIC:
    yoy = (p["revenue"] / p["revenue_prior"] - 1) if p["revenue_prior"] else None
    comp.append([p["name"], "Public", f"{yoy:+.1%} YoY (rev ${p['revenue']:.0f}M)",
                 f"{p['gross_margin']:.0%}", f"{p['rnd']/p['revenue']:.0%}",
                 f"net margin {p['net_margin']:+.0%}",
                 "EV/Revenue requires live market cap; growth/margins from FY2024 filings"])
# private capital efficiency = total disclosed raised / employees (funding-per-employee)
from collections import defaultdict
raised = defaultdict(float); emp = {}
for p in PRIVATE:
    if p["amount_musd"]:
        raised[p["name"]] += p["amount_musd"]
    emp[p["name"]] = p.get("employees")
for name in dict.fromkeys(p["name"] for p in PRIVATE):
    e = emp.get(name); tot = raised.get(name, 0)
    fpe = (tot / e) if e else None
    last_val = next((p["post_money_busd"] for p in reversed(PRIVATE) if p["name"] == name and p["post_money_busd"]), None)
    comp.append([name, "Private",
                 f"last val ${last_val:.1f}B" if last_val else "val n/d",
                 "n/d (private)", "n/d (private)",
                 f"${fpe:.1f}M raised/employee" if fpe else "n/d",
                 "Funding-per-employee is a rough proxy; headcounts approximate, rounds may be partial"])
_rows(ws, comp, flag_col=7)
for c, w in enumerate([16, 9, 26, 13, 13, 24, 50], 1):
    ws.column_dimensions[gl(c)].width = w
ws.freeze_panes = "A3"

# ---------------- summary ----------------
ws = wb.create_sheet("summary"); ws.sheet_view.showGridLines = False
_title(ws, "Summary — peer benchmark read for InGen", 2)
notes = [
 ("Public scale leaders", "Teradyne (~$2.8B rev, 58% GM, 19% net margin) and Symbotic (~$1.8B rev, hardware-thin margins) anchor the large-cap end; iRobot (~$0.68B, 21% GM, loss-making) and Cognex (~$0.92B, ~68% GM) are mid-scale."),
 ("Margin structure", "Vision/test peers (Cognex ~68%, Teradyne ~58% GM) are far more profitable than robotics-hardware peers (iRobot 21%, Symbotic ~18% GM). Hardware-heavy robotics carries structurally lower gross margin — relevant for InGen's own unit economics."),
 ("R&D intensity", "iRobot ~14% and Cognex ~16% of revenue on R&D even while pressured on margin; Teradyne ~16%. High R&D intensity is the norm in this space."),
 ("Private capital", "Humanoid privates are extraordinarily well-funded: Figure ($2.6B Series-B post, then $39B), Apptronik (~$1.5B), Agility ($2.12B). Capital raised >> current revenue — these are bets on future deployment, not current financials."),
 ("Capital efficiency caveat", "Funding-per-employee is high across humanoid privates (capital pouring in ahead of scaled commercial revenue). Treat valuations as momentum signals, not earnings-backed."),
 ("For InGen's narrative", "Benchmark on the RIGHT peer set per vertical: vision/test peers for margin structure; robotics-hardware peers for gross-margin reality; humanoid privates for the funding-vs-IP story (well-funded != defensible — see Week 2)."),
 ("Reconciliation note", "Public figures are FY2024 from SEC filings/earnings; cells marked 'approx/annualised' (yellow) should be confirmed against the exact 10-K/10-Q. Run src/financials/edgar_fetch.py (INGEST_ALLOW_NETWORK=1) for filing-exact 12-quarter history."),
 ("Sources", f"SEC EDGAR 10-K/10-Q & earnings releases (public); company press releases & reputable tech press (private). Retrieved {RETRIEVED}. Every row carries its source."),
]
ws.cell(row=2, column=1, value="Topic").font = HF; ws.cell(row=2, column=1).fill = PatternFill("solid", fgColor=BLUE); ws.cell(row=2, column=1).border = BD
ws.cell(row=2, column=2, value="Read").font = HF; ws.cell(row=2, column=2).fill = PatternFill("solid", fgColor=BLUE); ws.cell(row=2, column=2).border = BD
for i, (a, b) in enumerate(notes, 3):
    ws.cell(row=i, column=1, value=a).font = Font(bold=True, size=10, color=NAVY)
    ws.cell(row=i, column=2, value=b).font = CF
    for c in (1, 2):
        ws.cell(row=i, column=c).border = BD
        ws.cell(row=i, column=c).alignment = Alignment(vertical="top", wrap_text=True)
    ws.row_dimensions[i].height = 46
ws.column_dimensions["A"].width = 24; ws.column_dimensions["B"].width = 104

wb.save(OUT)
print(f"workbook -> {OUT.relative_to(ROOT)}")
if __name__ == "__main__":
    pass
